import cv2
import numpy as np
import face_recognition
import openpyxl
from keras.models import load_model
import pandas as pd
from datetime import datetime

# Load trained face recognition model
model = load_model('./Model/face_recognition_model')

# Load dataset information
dataset = pd.read_csv('./Dataset/Dataset.csv')

# Load attendance Excel sheet
wb = openpyxl.load_workbook('./attendance.xlsx')
sheet = wb.active

def mark_attendance(name):
    today = datetime.today().strftime('%Y-%m-%d')
    col = sheet.max_column
    if today not in [sheet.cell(1, i).value for i in range(2, col + 1)]:
        sheet.cell(1, col + 1).value = today
    col = sheet.max_column
    if name not in [sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1)]:
        sheet.cell(sheet.max_row + 1, 1).value = name
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row, 1).value == name:
            sheet.cell(row, col).value = 'Present'

def recognize_faces_in_image(image_path):
    # Load image
    image = cv2.imread(image_path)
    
    # Convert image to RGB (face_recognition library expects RGB format)
    rgb_image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    
    # Find face locations in the image
    face_locations = face_recognition.face_locations(rgb_image)
    if len(face_locations) == 0:
        print(f"No faces found in {image_path}")
        return

    # Extract face encodings
    face_encodings = face_recognition.face_encodings(rgb_image, face_locations)

    for face_encoding in face_encodings:
        # Predict using the trained model
        prediction = model.predict(np.expand_dims(face_encoding, axis=0))
        predicted_label = np.argmax(prediction)
        
        # Get the corresponding name from Dataset.csv
        name = dataset.loc[dataset['label'] == predicted_label, 'id'].values[0]
        
        mark_attendance(name)

def main():
    # List of test image paths
    test_images = ['./Dataset/Faces/Akshay Kumar_17.jpg', './Dataset/Faces/Alexandra Daddario_90.jpg', './Dataset/Faces/Zac Efron_89.jpg']  # Add paths to your test images

    # Process each test image
    for image_path in test_images:
        recognize_faces_in_image(image_path)

    # Save attendance
    wb.save('./attendance.xlsx')

if __name__ == "__main__":
    main()
